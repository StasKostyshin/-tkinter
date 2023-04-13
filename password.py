import pandas
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import numpy as np
from itertools import islice
#from colorama import init
#from colorama import Fore, Back, Style
import time
#import win32com.client as win32
from tkinter import *
import tkinter as tk 
from PIL import ImageTk, Image
import pickle
from tkinter import messagebox
#init()
##################################Регистрация#####################
def registration():
	winreg=tk.Toplevel()
	winreg.title('регистрирование пользователя')
	winreg.geometry("550x500+50+50")
	winreg.resizable(False,False)

	creg=Canvas(winreg,width=550, height=500, bg='gold')
	creg.pack(expand=True, fill= BOTH)

	creg.create_text(270,20,justify=CENTER,width=520, text='Регистрация',font=('Courier',20,'bold'))
	creg.create_text(270,70, text='Введите логин',font=('Courier',14, 'bold'))
	reg_login=Entry(creg, font=('Courier',14,'bold'))
	creg.create_window(170, 85, anchor=NW, window=reg_login, width=220, height=30 )
	creg.create_text(270,150, text='Введите пароль',font=('Courier',14, 'bold'))
	reg_password=Entry(creg, show="*", font=('Courier',14,'bold'))
	creg.create_window(170,170, anchor=NW, window=reg_password, width=220, height=30 )

	butreg=tk.Button(creg, text='Регистрация', font=('Courier',14,'bold'),command=lambda: [save()])
	creg.create_window(170, 220, anchor=NW, window=butreg, width=220, height=50)

	def save():
		login_pass_save={} #Запись в 
		login_pass_save[reg_login.get()]=reg_password.get()# Получаем введённые данные
		f = open('log.txt', 'wb') #Запись в файл
		pickle.dump(login_pass_save, f)   #Сохраняем в файл
		creg.create_text(270,310, text='Регистрация завершена, данные сохранены',font=('Courier',14, 'bold'))
		f.close()
		butreg=tk.Button(creg, text='Выход', font=('Courier',14,'bold'),command=lambda: [winreg.destroy()])
		creg.create_window(170, 370, anchor=NW, window=butreg, width=220, height=50)
#########################Ввод логина и пароля######################
def login():
	winlog=tk.Toplevel()
	winlog.title('Введите логин и пароль')
	winlog.geometry("550x500+50+50")
	winlog.resizable(False,False)

	clog= Canvas(winlog,width=550, height=500, bg='red')
	clog.pack(expand=True, fill= BOTH)

	clog.create_text(270,20,justify=CENTER,width=520, text='Авторизация',font=('Courier',20,'bold',))
	clog.create_text(270,70, text='Введите логин',font=('Courier',14, 'bold'))
	entry_login=Entry(clog, font=('Courier',14,'bold'))
	clog.create_window(170, 85, anchor=NW, window=entry_login, width=220, height=30 )
	clog.create_text(270,150, text='Введите пароль',font=('Courier',14, 'bold'))
	entry_password=Entry(clog, show="*", font=('Courier',14,'bold'))
	clog.create_window(170,170, anchor=NW, window=entry_password, width=220, height=30 )

	butlog=tk.Button(clog, text='Ввод', font=('Courier',14,'bold'),command=lambda: [log_pas(),winlog.destroy()])
	clog.create_window(170, 220, anchor=NW, window=butlog, width=220, height=50)

	def log_pas():
		f=open("log.txt", 'rb')
		a=pickle.load(f)
		f.close()
		if entry_login.get() in a:
			if entry_password.get()==a[entry_login.get()]:
				messagebox.showinfo('Всё верно','Можете продолжить')
				registration()
			if entry_password.get()!=a[entry_login.get()]:
				messagebox.showerror('Ошибка','Неверный логин или пароль')
			
		else:
			messagebox.showerror('Ошибка','Неверный логин')
		return

#######################Функция о программе#######################
def about():
	
	winabout=tk.Toplevel()
	winabout.title('О программе')
	winabout.geometry("550x500+50+50")
	winabout.resizable(False,False)
	cabout= Canvas(winabout,width=550, height=500, bg='gold')
	cabout.pack(expand=True, fill= BOTH)

	cabout.create_text(270,100,justify=CENTER,width=520, text='Программа созданна для улучшения работы сотрудников охраны, основанная на автопоиске информации по проходному режиму',font=('Courier',20,'bold'))
	cabout.create_text(60, 460, text='production \nKastyshyn S.M.',font=('Courier',8,'bold'))
##########################Функция о руководстве########################
def helpspravka():
	winhelp=tk.Toplevel()
	winhelp.title('Руководство')
	winhelp.geometry("550x500+50+50")
	winhelp.resizable(False,False)

	cabout= Canvas(winhelp,width=550, height=500, bg='gold')
	cabout.pack(expand=True, fill= BOTH)

	cabout.create_text(270,50,justify=CENTER,width=520, text='Руководство',font=('Courier',20,'bold'))


####################################################################
inavto = [""]
avto=0
inpeople = [""]
iloc= [ ]#Пустая переменая для строк из екселя


msg_input=" Заехал на территорию МЦК. Есть в списках. В:"
msg_output=" Выехал с МЦК. В:"
msg_inputnone=" Выехал но не заезжал на МЦК. В:"
msg_notfound=" В списках нет. Заехал на территорию МЦК. В:"


def protokol_write(str_write):#Функция протокол записи(строка ввода, указаны выше)
	pr=('Контролёр: '+str(entry_user.get()+str(iloc)+str(avto)+str_write+' '+time.ctime()+'\n'))
	with open("avtoinout.txt", "a") as a_writer:#открытие файла масив.тхт для 'a'-дозаписи
		a_writer.write(pr)#дозапись данных в переменную pr(указана выше)
		a_writer.close()
		

msg_input_people=" Зашёл на территорию МЦК. Есть в списках. В:"
msg_output_people=" Вышел с МЦК.В:"
msg_inputnone_people=" Вышел, но не заходил на МЦК. В:"
msg_notfound_people=" В списках нет. Зашёл на территорию МЦК. В:"


def protokol_write_people(str_write):#Функция протокол записи(строка ввода, указаны выше)
	pr=('Контролёр: '+str(entry_user.get()+str(iloc)+str(people)+str_write+' '+time.ctime()+'\n'))
	with open("peopleinout.txt", "a") as a_writer:#открытие файла масив.тхт для 'a'-дозаписи
		a_writer.write(pr)#дозапись данных в переменную pr(указана выше)
		a_writer.close()
		

def excel():
	df = pd.DataFrame({'Организации':[0],'Номер':[0]})#Пока единственный рабочий
	df.to_excel("Avtospisok.xlsx", index=False)#Создаём ексель файл

def read_excel():
	os.startfile("Avtospisok.xlsx")#Открытие Эксель файла

def sobitie():
	os.startfile("avtoinout.txt")#Открые файла тхт

def excel_people():
	df = pd.DataFrame({'Организации':[0],'Сотрудник':[0]})#Пока единственный рабочий
	df.to_excel("peoplespisok.xlsx", index=False)#Создаём ексель файл

def read_excel_people():
	os.startfile("peoplespisok.xlsx")#Открытие Эксель файла

def sobitie_people():
	os.startfile("peopleinout.txt")

###########################Выбрать картинку#####################
def workstart():
	win = tk.Toplevel()
	win.title('Въезд.Выезд')
	win.geometry("550x500+50+50")
	win.resizable(False, False)
	
	c9 = Canvas(win, width=550, height=500, bg='blue')
	c9.pack(expand=True, fill= BOTH)
	c9.create_image(-40,-20, image=bg5, anchor="nw")

	c9.create_text(290, 20,font=('Courier',14,'bold'),text="Выберите Въезд либо Выезд")

	
	btn1=tk.Button(win, text='Въезд',font=('Courier',14,'bold'),command=lambda:[proverka(), win.destroy()])
	c9.create_window(170,40, anchor=NW, window=btn1, width=220, height=30)
	btn2=tk.Button(win, text='Выезд',font=('Courier',14,'bold'),command=lambda:[proverka2(), win.destroy()])
	c9.create_window(170,80, anchor=NW, window=btn2, width=220, height=30)

def proverka():
	win2 = tk.Toplevel()
	win2.title('Ввод Въезда')
	win2.geometry("550x500+50+50")
	win2.resizable(False, False)
	
	c8 = Canvas(win2, width=550, height=500, bg='blue')
	c8.pack(expand=True, fill= BOTH)
	c8.create_image(-40,-20, image=bg5, anchor="nw")

	c8.create_text(270, 20,font=('Courier',14,'bold'),text="Введите номер авто в формате \"xxxx-y\"",)

	global entry_avto
	entry_avto=Entry(c8, font=('Courier',14))
	c8.create_window(170, 40, anchor=NW, window=entry_avto, width=220, height=30)
	btn=tk.Button(c8, text='Ввод',font=('Courier',14,'bold'),command=lambda:[avtoin(), win2.destroy()])
	c8.create_window(170,80, anchor=NW, window=btn, width=220, height=30)
###########################Найти Картинку 
def avtoin():
	df = pd.read_excel("Avtospisok.xlsx")
	wb = load_workbook("Avtospisok.xlsx")#Загрузка из Эксел в рабочую книгу
	sheet = wb['Sheet1']#Указываем с каким листом работаем в данном случае это лист'Тачки' 
	obl=list(df.iloc[0:200])#Рабочий
	def click():
		btn=tk.Button(c7,text='Продолжить',font=('Courier',14,'bold'),command=lambda: [workstart(), win3.destroy()])
		c7.create_window(170,130, anchor=NW, window=btn, width=220, height=30)
		btn1=tk.Button(c7,text='Нет',font=('Courier',14,'bold'),command=lambda: [win3.destroy()])
		c7.create_window(170,170, anchor=NW, window=btn1, width=220, height=30)
		
	win3 = tk.Toplevel()
	win3.title('Въезд')
	win3.geometry("550x500+50+50")
	win3.resizable(False, False)

	c7 = Canvas(win3, width=550, height=500, bg='blue')
	c7.pack(expand=True, fill= BOTH)
	c7.create_image(-40,-20, image=bg5, anchor="nw")

	global avto
	avto = 0
	avto=entry_avto.get()
	inavto.append(avto)
	found = False#если не нашёл введённые данные перепрыгивает на if not found
	for obl in range(sheet.max_row-1):#Диапозон поиска максимум строк - 1 первая ячейка 0 из 0
		if avto in list(df.iloc[obl]):#Если введённые данные авто находятся в строке(df.iloc[obl-диапазон указаный выше]) 
			found = True#Если введённые данные были были найденные
			c7.create_text(275, 60,justify=CENTER,width=520, font=('Courier',14,'bold'), text=(str(avto)+" Есть в списках: "'\n'+ df.iloc[obl,0]+'\n'"Дата и время: ", time.ctime()))
			iloc.append(df.iloc[obl,0])  
			protokol_write(msg_input)
			iloc.remove(df.iloc[obl,0])
			
	if not found:# Если введённые данные не найдены производит действия после :  
		c7.create_text(275,40,justify=CENTER,width=520, font=('Courier',14,'bold'), text=(str(avto)+" Нет в списках", time.ctime()))
		protokol_write(msg_notfound)
	click()
	return

def proverka2():
	winpr1 = tk.Toplevel()
	winpr1.title('Ввод Выезда')
	winpr1.geometry("550x500+50+50")
	winpr1.resizable(False, False)
	
	c6 = Canvas(winpr1, width=550, height=500, bg='blue')
	c6.pack(expand=True, fill= BOTH)
	c6.create_image(-40,-20, image=bg5, anchor="nw")

	global entry_avto
	c6.create_text(270,20, font=('Courier',14,'bold'), text="Введите номер авто в формате \"xxxx-y\"")
	
	entry_avto=Entry(c6, font=('Courier',14,'bold'))
	c6.create_window(170, 40, anchor=NW, window=entry_avto, width=220, height=30)
	bun=tk.Button(c6, text='Ввод',font=('Courier',14,'bold'),command=lambda: [avtoout(), winpr1.destroy()])
	c6.create_window(170,80, anchor=NW, window=bun, width=220, height=30)


def avtoout():
	def click():
		btn=tk.Button(c5,text='Продолжить',font=('Courier',14,'bold'),command=lambda: [workstart(), win3.destroy()])
		c5.create_window(170,90, anchor=NW, window=btn, width=220, height=30)
		btn1=tk.Button(c5,text='Нет',font=('Courier',14,'bold'),command=lambda: [win3.destroy()])
		c5.create_window(170,130, anchor=NW, window=btn1, width=220, height=30)
	win3 = tk.Toplevel()
	win3.title('Выезд')
	win3.geometry("550x500+50+50")
	win3.resizable(False, False)
	
	c5 = Canvas(win3, width=550, height=500, bg='grey')
	c5.pack(expand=True, fill= BOTH)
	c5.create_image(-40,-20, image=bg4, anchor="nw")

	global avto
	avto=entry_avto.get()	
	if avto in inavto:
		c5.create_text(260,40,justify=CENTER,font=('Courier',14,'bold'), text=(str(avto)+" Выезжает", time.ctime()))
		protokol_write(msg_output)
		inavto.remove(avto)
		click()
	else:
		c5.create_text(260,40,justify=CENTER,font=('Courier',14,'bold'), text=(str(avto)+" Не заезжала", time.ctime()))
		protokol_write(msg_inputnone)
		click()
################Функции для столовых машин############################
def prosmotr():
	os.startfile("Stolovaia.xlsx")

def excelstolovaia():
	df = pd.DataFrame({'Дата':[0],'Контролёр':[0],'Номер Авто':[0],'Фамилия':[0],'Организация':[0],'Номер ТТН':[0]})#Пока единственный рабочий
	df.to_excel("Stolovaia.xlsx", index=False)#Создаём ексель файл 

def  vvod():
	en=entry_user.get()#при помощи get()перенимаем переменную из entry_user
	number=entry_number.get()#при помощи get()перенимаем переменную из entry_number
	fio=entry_fio.get()#при помощи get()перенимаем переменную из entry_fio
	organization=entry_organization.get()#при помощи get()перенимаем переменную из entry_organization
	ttn=entry_ttn.get()#при помощи get()перенимаем переменную из entry_ttn

	wb = load_workbook("Stolovaia.xlsx")
	sheet = wb['Sheet1']
	sheet.append([time.ctime(),en,number,fio,organization,ttn])
	wb.save("Stolovaia.xlsx")
	wb.close()

	winvvod = tk.Toplevel()#Создаём окно
	winvvod.title('Сохранение в файл')#Даём название окну
	winvvod.geometry("550x500+50+50")#Задаём размеры окна
	winvvod.resizable(False, False)#Запрет на изменение окна по Х,Y

	c4 = Canvas(winvvod, width=550, height=500, bg='grey')
	c4.pack(expand=True, fill= BOTH)
	c4.create_image(0,0, image=bg3, anchor="nw")

	c4.create_text(270,20, text='Данные сохранены',font=('Courier',20,'bold'))

	but=tk.Button(c4, text='Продолжить', font=('Courier',14),command=lambda: [winvvod.destroy()])
	c4.create_window(170,50, anchor=NW, window=but, width=220, height=30)
	
####################Найти катинку другую###################	
def start():
	winstart = tk.Toplevel()#Создаём окно
	winstart.title('Данные ТТН')#Даём название окну
	winstart.geometry("550x500+50+50")#Задаём размеры окна

	winstart.resizable(False, False)#Запрет на изменение окна по Х,Y
	c3 = Canvas(winstart, width=550, height=500, bg='green')
	c3.pack(expand=True, fill= BOTH)
	c3.create_image(0,0, image=bg2, anchor="nw")
	
	c3.create_text(270,20, text='Введите танные с ТТН',font=('Courier',20,'bold'))

	global entry_number
	global entry_fio
	global entry_organization
	global entry_ttn
	c3.create_text(270,60, text='Номер авто',font=('Courier',14, 'bold'))
	entry_number=Entry(c3, font=('Courier',14,'bold'))
	c3.create_window(170, 70, anchor=NW, window=entry_number, width=220, height=30 )
	
	c3.create_text(270,120, text='Фамилия',font=('Courier',14, 'bold'))
	entry_fio=Entry(c3, font=('Courier',14,'bold'))
	c3.create_window(170, 130, anchor=NW, window=entry_fio, width=220, height=30 )

	c3.create_text(270,180, text='Организация',font=('Courier',14, 'bold'))
	entry_organization=Entry(c3, font=('Courier',14,'bold'))
	c3.create_window(170,190, anchor=NW, window=entry_organization, width=220, height=30 )
	
	c3.create_text(270,240, text='Номер ТТН',font=('Courier',14, 'bold'))
	entry_ttn=Entry(c3, font=('Courier',14,'bold'))
	c3.create_window(170, 250, anchor=NW, window=entry_ttn, width=220, height=30 )

	but=tk.Button(c3, text='Сохранить', font=('Courier',14,'bold'),command=lambda: [vvod(), winstart.destroy()])
	c3.create_window(160,310, anchor=NW, window=but, width=240, height=30)
	
#####Создать главное окно столовой########
def stolovaia():
	winst = tk.Toplevel()#Создаём окно
	winst.title('Авто в столовую')#Даём название окну
	winst.geometry("550x500+50+50")#Задаём размеры окна

	winst.resizable(False, False)#Запрет на изменение окна по Х,Y
	c2 = Canvas(winst, width=550, height=500, bg='green')
	c2.pack(expand=True, fill= BOTH)#expand-помещение в нижней части котейнера, fiil-Both заполняет пространство по Х и У
	c2.create_image(0,0, image=bg2, anchor="nw")#Устанавливаем картинку 0-х.0-y, картинка задний фон, расположение=Nort,West

	c2.create_text(270,20, text='Выберите действие',font=('Courier',20,'bold'))#Выводим текс на Экран

	but1=tk.Button(c2, text='Ввести данные с ТТН', font=('Courier',14,'bold'),command=start)
	c2.create_window(170, 40, anchor=NW, window=but1, width=220, height=50)
	but2=tk.Button(c2, text='Просмотреть журнал', font=('Courier',14,'bold'), command=prosmotr)
	c2.create_window(170, 110, anchor=NW, window=but2, width=220, height=50)
	but3=tk.Button(c2, text='Выход', font=('Courier',14,'bold'), command=lambda: [winst.destroy()])
	c2.create_window(170, 170, anchor=NW, window=but3, width=220, height=50)

	but4=tk.Button(c2, text='Создать журнал/Очистить', font=('Courier',14,'bold'), command=lambda: [excelstolovaia(),c2.create_text(270, 330, text='Файл создан',font=('Arial',14,'bold'), fill='red')],bg='red')
	c2.create_window(120, 230, anchor=NW, window=but4, width=320, height=50)
###########Создаём окно вьезда по спискам#########
def spiski():
	winsp = tk.Toplevel()#Создаём окно
	winsp.title('По годовым спискам авто')#Даём название окну
	winsp.geometry("550x500+50+50")#Задаём размеры окна
	winsp.resizable(False, False)
	c1 = Canvas(winsp, width=550, height=500)
	c1.pack(expand=True, fill= BOTH)
	c1.create_image(-20,0, image=bg1, anchor="nw")
	
	c1.create_text(270,20, text='Выберите действие!',font=('Courier',20,'bold'))#Выводим текс на Экран

	btn4=tk.Button(c1, text='Создать',font=('Courier',14,'bold'),command=lambda: [excel(), c1.create_text(300, 40, text='Файл создан',font=('Arial',14,'bold'), fill='red')],bg='red')
	c1.create_window(30, 40, anchor=NW, window=btn4, width=120, height=50)
	
	btn5=tk.Button(c1, text='Изменить',font=('Courier',14,'bold'),command=read_excel)#Устанавливаем кнопку 1
	c1.create_window(400, 40, anchor=NW, window=btn5, width=120, height=50)

	btn6=tk.Button(c1, text='Просмотр событий',font=('Courier',14,'bold'),command=sobitie)#Устанавливаем кнопку 1
	c1.create_window(170, 100, anchor=NW, window=btn6, width=220, height=50)
	
	btn7=tk.Button(c1, text='Работать с программой',font=('Courier',14,'bold'),command=lambda: [workstart(), winsp.destroy()])
	c1.create_window(160, 160, anchor=NW, window=btn7, width=240, height=50)

	btn8=tk.Button(c1, text='Выход',font=('Courier',14,'bold'),command=lambda: [winsp.destroy()])
	c1.create_window(160, 220, anchor=NW, window=btn8, width=240, height=50)

#########################################Люди список#######################
def peoplemain():
	winsp = tk.Toplevel()#Создаём окно
	winsp.title('По годовым спискам посетителей')#Даём название окну
	winsp.geometry("550x500+50+50")#Задаём размеры окна
	winsp.resizable(False, False)
	c1 = Canvas(winsp, width=550, height=500, bg='yellow')
	c1.pack(expand=True, fill= BOTH)
	c1.create_image(0,0, image=bg6, anchor="nw")
	
	c1.create_text(270,20, text='Выберите действие!',font=('Courier',20,'bold'))#Выводим текс на Экран

	btn4=tk.Button(c1, text='Создать',font=('Courier',14,'bold'),command=lambda: [excel_people(), c1.create_text(300, 40, text='Файл создан',font=('Arial',14,'bold'), fill='red')],bg='red')
	c1.create_window(30, 40, anchor=NW, window=btn4, width=120, height=50)
	
	btn5=tk.Button(c1, text='Изменить',font=('Courier',14,'bold'),command=read_excel_people)#Устанавливаем кнопку 1
	c1.create_window(400, 40, anchor=NW, window=btn5, width=120, height=50)

	btn6=tk.Button(c1, text='Просмотр событий',font=('Courier',14,'bold'),command=sobitie_people)#Устанавливаем кнопку 1
	c1.create_window(170, 100, anchor=NW, window=btn6, width=220, height=50)
	
	btn7=tk.Button(c1, text='Работать с программой',font=('Courier',14,'bold'),command=lambda: [peopleinout(), winsp.destroy()])
	c1.create_window(160, 160, anchor=NW, window=btn7, width=240, height=50)

	btn8=tk.Button(c1, text='Выход',font=('Courier',14,'bold'),command=lambda: [winsp.destroy()])
	c1.create_window(160, 220, anchor=NW, window=btn8, width=240, height=50)

	btn7=tk.Button(c1, text='Просмотр посетителей',font=('Courier',14,'bold'),command=lambda: [lockinpeople(), winsp.destroy()])
	c1.create_window(160, 280, anchor=NW, window=btn7, width=240, height=50)


def peopleinout():
	win = tk.Toplevel()
	win.title('Вход,Выход')
	win.geometry("550x500+50+50")
	win.resizable(False, False)
	
	c9 = Canvas(win, width=550, height=500, bg='yellow')
	c9.pack(expand=True, fill= BOTH)
	c9.create_image(0,0, image=bg6, anchor="nw")

	c9.create_text(290, 20,font=('Courier',14,'bold'),text="Выберите Вход либо Выход")

	
	btn1=tk.Button(win, text='Вход',font=('Courier',14,'bold'),command=lambda:[peoplein(), win.destroy()])
	c9.create_window(170,40, anchor=NW, window=btn1, width=220, height=30)
	btn2=tk.Button(win, text='Выход',font=('Courier',14,'bold'),command=lambda:[peopleout(), win.destroy()])
	c9.create_window(170,80, anchor=NW, window=btn2, width=220, height=30)
##################################Функции входа людей#########################
def peoplein():
	win2 = tk.Toplevel()
	win2.title('Ввод Входа')
	win2.geometry("550x500+50+50")
	win2.resizable(False, False)
	
	c8 = Canvas(win2, width=550, height=500, bg='yellow')
	c8.pack(expand=True, fill= BOTH)
	c8.create_image(0,0, image=bg6, anchor="nw")

	c8.create_text(270, 20,font=('Courier',14,'bold'),text="Введите посетителя в формате\"Иванов И.И.\"",)

	global entry_people
	entry_people=Entry(c8, font=('Courier',14))
	c8.create_window(170, 40, anchor=NW, window=entry_people, width=220, height=30)
	btn=tk.Button(c8, text='Ввод',font=('Courier',14,'bold'),command=lambda:[peopleinmck(), win2.destroy()])
	c8.create_window(170,80, anchor=NW, window=btn, width=220, height=30)

def peopleinmck():
	df = pd.read_excel("peoplespisok.xlsx")#Читаем файл эксель
	wb = load_workbook("peoplespisok.xlsx")#Загрузка из Эксел в рабочую книгу
	sheet = wb['Sheet1']#Указываем с каким листом работаем в данном случае это лист'Тачки' 
	obl=list(df.iloc[0:200])#Рабочий
	def click():
		btn=tk.Button(c7,text='Продолжить',font=('Courier',14,'bold'),command=lambda: [peopleinout(), win3.destroy()])
		c7.create_window(170,130, anchor=NW, window=btn, width=220, height=30)
		btn1=tk.Button(c7,text='Нет',font=('Courier',14,'bold'),command=lambda: [win3.destroy()])
		c7.create_window(170,170, anchor=NW, window=btn1, width=220, height=30)
		
	win3 = tk.Toplevel()
	win3.title('Вход')
	win3.geometry("550x500+50+50")
	win3.resizable(False, False)

	c7 = Canvas(win3, width=550, height=500, bg='yellow')
	c7.pack(expand=True, fill= BOTH)
	c7.create_image(0,0, image=bg6, anchor="nw")

	global people
	people = 0
	people=entry_people.get()
	inpeople.append(people)
	found = False#если не нашёл введённые данные перепрыгивает на if not found
	for obl in range(sheet.max_row-1):#Диапозон поиска максимум строк - 1 первая ячейка 0 из 0
		if people in list(df.iloc[obl]):#Если введённые данные авто находятся в строке(df.iloc[obl-диапазон указаный выше]) 
			found = True#Если введённые данные были были найденные
			c7.create_text(275, 60,justify=CENTER,width=520, font=('Courier',14,'bold'), text=(str(people)+" Есть в списках: "'\n'+ df.iloc[obl,0]+'\n'"Дата и время: ", time.ctime()))
			iloc.append(df.iloc[obl,0])  
			protokol_write_people(msg_input_people)
			iloc.remove(df.iloc[obl,0])
			
	if not found:# Если введённые данные не найдены производит действия после :  
		c7.create_text(275,60,justify=CENTER,width=520, font=('Courier',14,'bold'), text=(str(people)+" Нет в списках", time.ctime()))
		protokol_write_people(msg_notfound_people)
	click()
	return

def lockinpeople():
	win3 = tk.Toplevel()
	win3.title('Просмотр посетителей')
	win3.geometry("550x500+50+50")
	win3.resizable(False, False)

	c8 = Canvas(win3, width=550, height=500, bg='yellow')
	c8.pack(expand=True, fill= BOTH)
	c8.create_image(0,0, image=bg6, anchor="nw")
	

	c8.create_text(275,50,justify=CENTER,width=300,font=('Courier',12,'bold'),text=(str(inpeople)))
	btn=tk.Button(win3,text='Продолжить',font=('Courier',14,'bold'),command=lambda: [win3.destroy()])
	c8.create_window(170,170, anchor=NW, window=btn, width=220, height=30)



############################Функция выхода людей#########################
def peopleout():
	winpr1 = tk.Toplevel()
	winpr1.title('Ввод Выхода')
	winpr1.geometry("550x500+50+50")
	winpr1.resizable(False, False)
	
	c6 = Canvas(winpr1, width=550, height=500, bg='yellow')
	c6.pack(expand=True, fill= BOTH)
	c6.create_image(0,0, image=bg7, anchor="nw")

	global entry_people
	c6.create_text(270,20, font=('Courier',14,'bold'), text="Введите посетителя в формате\"Иванов И.И.\"")
	
	entry_people=Entry(c6, font=('Courier',14,'bold'))
	c6.create_window(170, 40, anchor=NW, window=entry_people, width=220, height=30)
	bun=tk.Button(c6, text='Ввод',font=('Courier',14,'bold'),command=lambda: [peopleout2(), winpr1.destroy()])
	c6.create_window(170,80, anchor=NW, window=bun, width=220, height=30)

def peopleout2():
	def click():
		btn=tk.Button(c5,text='Продолжить',font=('Courier',14,'bold'),command=lambda: [peopleinout(), win3.destroy()])
		c5.create_window(170,90, anchor=NW, window=btn, width=220, height=30)
		btn1=tk.Button(c5,text='Нет',font=('Courier',14,'bold'),command=lambda: [win3.destroy()])
		c5.create_window(170,130, anchor=NW, window=btn1, width=220, height=30)
	win3 = tk.Toplevel()
	win3.title('Выход')
	win3.geometry("550x500+50+50")
	win3.resizable(False, False)
	
	c5 = Canvas(win3, width=550, height=500, bg='yellow')
	c5.pack(expand=True, fill= BOTH)
	c5.create_image(0,0, image=bg7, anchor="nw")

	global people
	people=entry_people.get()	
	if people in inpeople:
		c5.create_text(260,40,justify=CENTER,width=520,font=('Courier',14,'bold'), text=(str(people)+" Выходит", time.ctime()))
		protokol_write_people(msg_output_people)
		inpeople.remove(people)
		click()
	else:
		c5.create_text(260,40,justify=CENTER,width=520, font=('Courier',14,'bold'), text=(str(people)+" Не заходил"'\n', time.ctime()))
		protokol_write_people(msg_inputnone_people)
		click()

#####################Первое окно##########################################
win=tk.Tk()
win.title('Транспортная')#Даём название окну
win.geometry("550x480+50+50")#Задаём размеры окна
win.iconphoto(True, tk.PhotoImage(file='znak-radiaciya.png'))#Устанавливаем иконку
win.resizable(False, False)
bg=ImageTk.PhotoImage(file='start.png')
bg1=ImageTk.PhotoImage(file='spiski.png')
bg2=ImageTk.PhotoImage(file='stolovaia.png')
bg3=ImageTk.PhotoImage(file='minon.png')
bg4=ImageTk.PhotoImage(file='mramor.png')
bg5=ImageTk.PhotoImage(file='plamia.png')
bg6=ImageTk.PhotoImage(file='peplymain.png')
bg7=ImageTk.PhotoImage(file='peaplyin.png')

c = Canvas(win, width=550, height=480, bg='red')
c.pack(expand=True, fill= BOTH)

user=StringVar(value='Введите Ф.И.О. контролёра')
entry_user=Entry(c, font=('Courier',14,'bold'),textvariable=user)
c.create_window(120, 120, anchor=NW, window=entry_user, width=320, height=30 )
btn=tk.Button(text='Пользователь', font=('Courier',14,'bold'),command=lambda: [onbutten()])
c.create_window(170, 155, anchor=NW, window=btn, width=220, height=30)
c.create_text(60, 460, text='production \nKastyshyn S.M.',font=('Courier',8,'bold'))

c.create_image(-35,30, image=bg, anchor="nw")
win.config(bg='blue')#Устанавливаем задний фон
c.create_text(300, 20, text='С какими списками работать?',font=('Courier',20,'bold'))
btn1=tk.Button(text='По ТТН', font=('Courier',14,'bold'), command=lambda: [stolovaia()],state=DISABLED,bg='green')
c.create_window(20, 40,anchor=NW, window=btn1, width=120, height=50)	
btn2=tk.Button(text='По спискам', font=('Courier',14,'bold'), command=lambda: [spiski()], state=DISABLED,bg='blue')
c.create_window(400, 40, anchor=NW, window=btn2, width=120, height=50)	
btn3=tk.Button(text='Выход', font=('Courier',14,'bold'),command=lambda: [win.quit()],bg='red')
c.create_window(220, 200, anchor=NW, window=btn3, width=120, height=50)
btn4=tk.Button(text='Посетители', font=('Courier',14,'bold'), command=lambda: [peoplemain()], state=DISABLED,bg='yellow')
c.create_window(220, 40, anchor=NW, window=btn4, width=130, height=50)	
def onbutten():
	if btn['state'] =='normal':
		btn['state']='disabled'
		btn1['state'] ='normal'
		btn2['state'] ='normal'
		btn4['state'] ='normal'
		entry_user['state'] = 'disabled'

m=Menu(win)
win.config(menu=m)

fm = Menu(m, tearoff=0)
m.add_cascade(label='Администрирование', menu=fm)
fm.add_command(label='Регистрация', command=login)

hm = Menu(m, tearoff=0)
m.add_cascade(label='Помощь', menu=hm)
hm.add_command(label='Справка', command=helpspravka)
hm.add_command(label='О программе', command=about)

win.mainloop()